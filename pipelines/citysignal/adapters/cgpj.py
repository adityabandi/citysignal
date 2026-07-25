"""CGPJ — mortgage foreclosures and evictions, the courts' own count of distress.

The Consejo General del Poder Judicial publishes "Efectos de la crisis en los
órganos judiciales," a quarterly statistical series born out of the 2008
financial crisis and never discontinued, because the two figures it tracks never
stopped mattering: ``ejecuciones hipotecarias`` (mortgage foreclosure
proceedings filed with a court) and ``lanzamientos`` (evictions actually carried
out — the sheriff at the door, not just a court order). These are downstream of
almost everything else in the distress pack: a foreclosure is filed after months
of missed payments, and an eviction happens after the foreclosure has already
run its course, so both lag the cycle by design rather than by accident. CGPJ
counts every judicial district (partido judicial) and reports by province and
by Comunidad Autónoma; this adapter keeps the province level, which is what
CitySignal's eight-city geography actually resolves to.

The CGPJ statistics portal does not expose an API. It publishes one Excel
workbook per quarter, buried under a "Temas > Estadística Judicial" page whose
filenames are neither stable nor predictable (they carry the current quarter's
label and, not infrequently, a "_revisado" suffix when a prior quarter's number
is corrected). Rather than guess a filename, this adapter reads the crisis-
statistics listing page and follows whichever link is titled "por provincias" —
but it also publishes a second workbook, one that is not quarter-stamped at all:
"Series - Efecto de la crisis en los órganos judiciales por provincias," a
running series from 2007 to the present quarter, rebuilt in place every time
CGPJ adds a quarter. Reading that one series file instead of forty-odd quarterly
snapshots turns nine years of history into a single request.
"""

from __future__ import annotations

import re
import urllib.parse
from io import BytesIO
from typing import Iterable

import openpyxl
import pandas as pd
from bs4 import BeautifulSoup

from ..framework.adapter import AdapterFailure, BaseAdapter, RunContext, SourceManifest
from ..framework.fetch import FetchPlan, RawPayload
from ..framework.record import CanonicalRecord, province

LISTING_URL = (
    "https://www.poderjudicial.es/cgpj/es/Temas/Estadistica-Judicial/"
    "Estadistica-por-temas/Datos-penales--civiles-y-laborales/Civil-y-laboral/"
    "Efecto-de-la-Crisis-en-los-organos-judiciales/"
)

# (sheet name, metric_id, unit) for the two blocks we need out of the workbook.
# Sheet names carry CGPJ's own stray leading/trailing spaces — matched verbatim.
SHEETS = (
    (" Ej.Hipot por provincias", "foreclosures", "cases"),
    ("Lanzamientos pract. Total prov", "evictions", "cases"),
)

# CGPJ's province names for our eight target provinces, as printed in column B
# of the series sheets (upper-case, unaccented where CGPJ itself omits accents).
PROVINCE_NAMES = {
    "MADRID": "28",
    "BARCELONA": "08",
    "VALENCIA": "46",
    "MALAGA": "29",
    "SEVILLA": "41",
    "ILLES BALEARS": "07",
    "BIZKAIA": "48",
    "ZARAGOZA": "50",
}

_QUARTER_RE = re.compile(r"^(\d{2})-T([1-4])$")


def _period_from_header(value: object) -> str | None:
    if not isinstance(value, str):
        return None
    match = _QUARTER_RE.match(value.strip())
    if not match:
        return None
    yy, q = match.groups()
    return f"{2000 + int(yy):04d}-Q{q}"


def _find_header_row(ws, max_scan: int = 15) -> int:
    for r in range(1, max_scan + 1):
        row = ws[r]
        hits = sum(1 for cell in row if _period_from_header(cell.value))
        if hits >= 4:
            return r
    raise AdapterFailure(f"could not locate a quarter header row in sheet {ws.title!r}")


def _extract_block(ws, metric_id: str, unit: str) -> list[dict[str, object]]:
    header_row = _find_header_row(ws)
    header = [c.value for c in ws[header_row]]
    quarter_cols = {idx: p for idx, h in enumerate(header) if (p := _period_from_header(h))}

    rows: list[dict[str, object]] = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        name = row[1] if len(row) > 1 else None
        if not name:
            continue
        name_clean = str(name).strip().upper()
        if name_clean == "TOTAL":
            break  # end of the absolute-count block; a % variación block follows
        ine_prov = PROVINCE_NAMES.get(name_clean)
        if not ine_prov:
            continue
        geo = province(ine_prov)
        for idx, period in quarter_cols.items():
            value = row[idx] if idx < len(row) else None
            if not isinstance(value, (int, float)):
                continue
            rows.append({"metric_id": metric_id, "geo_id": geo, "period": period, "value": float(value), "unit": unit})
    return rows


class CgpjAdapter(BaseAdapter):
    manifest = SourceManifest(
        source_id="cgpj",
        publisher="Consejo General del Poder Judicial",
        license="Reuse permitted (attribution)",
        attribution="Source: CGPJ, Estadística Judicial",
        docs_url="https://www.poderjudicial.es/cgpj/es/Temas/Estadistica-Judicial/",
        cadence="quarterly",
        geo_level="province",
        max_age_days=210,
        formats=("xlsx",),
        kind="official",
        redistribute=True,
        # CGPJ periodically republishes a quarter as "_revisado" with a corrected count.
        revisions_allowed=True,
        notes=(
            "Ejecuciones hipotecarias (foreclosures filed) and lanzamientos "
            "(evictions carried out) by province, from CGPJ's running series "
            "workbook rather than the forty-odd quarterly snapshot files."
        ),
    )

    def discover(self, ctx: RunContext) -> list[FetchPlan]:
        listing_plan = FetchPlan(url=LISTING_URL, fmt="html", label="cgpj-crisis-listing")
        try:
            payload = ctx.fetcher.get(listing_plan)
        except Exception as exc:  # noqa: BLE001
            raise AdapterFailure(f"could not fetch the CGPJ crisis-statistics listing page: {exc}") from exc
        if payload is None:
            raise AdapterFailure("CGPJ listing page fetch returned no content")

        href = self._find_series_link(payload.text())
        if not href:
            raise AdapterFailure(
                "no 'por provincias' series workbook link found on the CGPJ crisis-statistics page "
                "— the page layout may have changed"
            )

        absolute = urllib.parse.urljoin(LISTING_URL, href)
        parts = urllib.parse.urlsplit(absolute)
        safe_path = urllib.parse.quote(parts.path, safe="/")
        url = urllib.parse.urlunsplit((parts.scheme, parts.netloc, safe_path, parts.query, parts.fragment))

        return [FetchPlan(url=url, fmt="xlsx", label="series-por-provincias")]

    @staticmethod
    def _find_series_link(html: str) -> str | None:
        soup = BeautifulSoup(html, "lxml")
        for a in soup.find_all("a", href=True):
            href = a["href"]
            label = f"{href} {a.get('title', '')}".lower()
            if (
                href.lower().split("?")[0].endswith(".xlsx")
                and "por provincias" in label
                and "efecto de la crisis" in label
                and "tsj" not in label
            ):
                return href
        return None

    def parse(self, payload: RawPayload, ctx: RunContext) -> pd.DataFrame:
        workbook = openpyxl.load_workbook(BytesIO(payload.content), data_only=True, read_only=True)
        rows: list[dict[str, object]] = []
        for sheet_name, metric_id, unit in SHEETS:
            if sheet_name not in workbook.sheetnames:
                raise AdapterFailure(
                    f"expected sheet {sheet_name!r} not found in CGPJ workbook "
                    f"(available: {workbook.sheetnames})"
                )
            rows.extend(_extract_block(workbook[sheet_name], metric_id, unit))

        if not rows:
            raise AdapterFailure("parsed zero rows from the CGPJ series workbook")
        return pd.DataFrame(rows)

    def normalize(
        self, frame: pd.DataFrame, plan: FetchPlan, ctx: RunContext
    ) -> Iterable[CanonicalRecord]:
        for row in frame.itertuples():
            yield CanonicalRecord(
                metric_id=row.metric_id,
                geo_id=row.geo_id,
                period=row.period,
                value=float(row.value),
                unit=row.unit,
                source_id=self.manifest.source_id,
            )
