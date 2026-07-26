"""Building-permit visas (visados de dirección de obra) by province, monthly.

The Ministerio de Transportes y Movilidad Sostenible publishes construction
statistics including visados (building permits filed with the architecture board
before construction begins). These precede housing starts, which precede
completions by roughly 1-2 years, making them the earliest supply-side signal
in Spanish housing data. The series goes back decades.

The ministry publishes monthly statistics by province as XLS/XLSX files. This
adapter resolves the download URL by fetching the statistics listing page,
parsing it to find the Excel file link, then extracting province-level data.
"""

from __future__ import annotations

import io
import logging
import re
from typing import Any, Iterable

import pandas as pd
from bs4 import BeautifulSoup

from ..framework.adapter import AdapterFailure, BaseAdapter, RunContext, SourceManifest
from ..framework.fetch import FetchPlan, RawPayload
from ..framework.record import CanonicalRecord, province

log = logging.getLogger(__name__)

LISTING_URL = (
    "https://www.transportes.gob.es/informacion-para-el-ciudadano/"
    "informacion-estadistica/vivienda-y-actuaciones-urbanas"
)


def _num(value: Any) -> float | None:
    """Parse a value to float, handling Spanish number formats."""
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in ("nan", "null", "none", "_", "-"):
        return None
    try:
        # Handle both comma and period as decimal separators
        text = text.replace(".", "").replace(",", ".")
        return float(text)
    except ValueError:
        return None


def _prov_code(value: Any) -> str | None:
    """Convert a numeric province code to zero-padded 2-digit format."""
    n = _num(value)
    return None if n is None else f"{int(n):02d}"


class VisadosAdapter(BaseAdapter):
    manifest = SourceManifest(
        source_id="visados",
        publisher="Ministerio de Transportes y Movilidad Sostenible",
        license="Reuse permitted (attribution)",
        attribution="Source: Ministry of Transport, Building Permits Statistics",
        docs_url="https://www.transportes.gob.es/informacion-para-el-ciudadano/informacion-estadistica/vivienda-y-actuaciones-urbanas",
        cadence="monthly",
        geo_level="province",
        max_age_days=120,
        formats=("xlsx",),
        kind="official",
        min_rows=2,
        notes=(
            "Monthly building permits (visados de dirección de obra) filed with "
            "the architecture board, by province. The earliest signal of housing "
            "supply arriving, preceding starts by 1-2 years and completions by "
            "2-4 years. Series includes data back to the 1980s."
        ),
    )

    def discover(self, ctx: RunContext) -> list[FetchPlan]:
        """Fetch the ministry's statistics page and extract the visados Excel file URL."""
        listing_plan = FetchPlan(url=LISTING_URL, fmt="html", label="visados-listing")
        try:
            payload = ctx.fetcher.get(listing_plan)
        except Exception as exc:  # noqa: BLE001
            raise AdapterFailure(
                f"could not fetch the ministry's housing statistics listing page: {exc}"
            ) from exc
        if payload is None:
            raise AdapterFailure("Ministry listing page fetch returned no content")

        url = self._find_visados_link(payload.text())
        if not url:
            raise AdapterFailure(
                "no visados (building permits) Excel file link found on the ministry page "
                "— the page layout may have changed or the file may be temporarily unavailable"
            )

        return [FetchPlan(url=url, fmt="xlsx", label="visados-monthly")]

    @staticmethod
    def _find_visados_link(html: str) -> str | None:
        """Parse HTML to find the visados XLS/XLSX file URL."""
        soup = BeautifulSoup(html, "lxml")
        for a in soup.find_all("a", href=True):
            href = a["href"]
            text = f"{href} {a.get_text(), a.get('title', '')}".lower()
            # Look for links to Excel files with keywords suggesting visados/dirección/obra
            if (
                (href.lower().endswith(".xls") or href.lower().endswith(".xlsx"))
                and any(
                    kw in text
                    for kw in ("visado", "dirección", "obra", "edificios", "construcción")
                )
            ):
                return href
        return None

    def parse(self, payload: RawPayload, ctx: RunContext) -> pd.DataFrame:
        """Load the Excel file and extract the data table."""
        import openpyxl

        try:
            workbook = openpyxl.load_workbook(io.BytesIO(payload.content), data_only=True, read_only=True)
        except Exception as exc:  # noqa: BLE001
            raise AdapterFailure(f"failed to open Excel file: {exc}") from exc

        if not workbook.sheetnames:
            raise AdapterFailure("Excel file contains no sheets")

        # Try to find the data sheet (usually the first one)
        sheet = workbook[workbook.sheetnames[0]]

        # Extract rows, skipping header
        rows = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), 1):
            if row_idx == 1:
                # Skip header row or extract it
                continue
            if row and any(row):  # Skip completely empty rows
                rows.append(row)

        if not rows:
            raise AdapterFailure("no data rows found in Excel file")

        # Create DataFrame from rows
        # Assume columns are: Province, Year, Month, Value (or similar)
        df = pd.DataFrame(rows)
        return df

    def normalize(
        self, frame: pd.DataFrame, plan: FetchPlan, ctx: RunContext
    ) -> Iterable[CanonicalRecord]:
        """Convert DataFrame rows to canonical records."""
        prov_codes = ctx.config.province_codes

        for row in frame.itertuples(index=False):
            # Attempt to parse the row - columns vary by file structure
            # Try to infer: province code, year, month, value
            if len(row) < 3:
                continue

            # Try different column positions
            prov_code = None
            year = None
            month = None
            value = None

            # Common patterns:
            # Pattern 1: [province_code, year, month, ..., value]
            # Pattern 2: [province_name, year, month, value]
            # Pattern 3: [province_name, year, month, ..., multiple_values]

            # Try to find numeric patterns
            numeric_vals = []
            for cell in row:
                n = _num(cell)
                if n is not None:
                    numeric_vals.append(n)

            if len(numeric_vals) < 3:
                continue

            # Assume first numeric value could be province code (2-3 digits)
            if 1 <= numeric_vals[0] <= 52:
                prov_code = f"{int(numeric_vals[0]):02d}"
                year = int(numeric_vals[1])
                month = int(numeric_vals[2])
                value = numeric_vals[3] if len(numeric_vals) > 3 else numeric_vals[2]
            # Or assume second and third are year/month
            elif len(numeric_vals) >= 4:
                prov_code = f"{int(numeric_vals[0]):02d}"
                year = int(numeric_vals[1])
                month = int(numeric_vals[2])
                value = numeric_vals[3]
            else:
                continue

            if prov_code not in prov_codes:
                continue
            if not (1 <= month <= 12):
                continue
            if value is None or value < 0:
                continue

            yield CanonicalRecord(
                metric_id="visados_new_build",
                geo_id=province(prov_code),
                period=f"{year:04d}-{month:02d}",
                value=value,
                unit="dwellings",
                source_id=self.manifest.source_id,
            )
