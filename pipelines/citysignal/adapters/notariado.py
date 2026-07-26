"""Housing sales and prices from notary records (CIEN — Centro de Información Estadística del Notariado).

The Consejo General del Notariado publishes CIEN, a monthly statistical series
on housing sales transactions and median prices per square metre, by province.
Because notaries see the sale at signing (before the land registry records it),
notary statistics lead the registry by approximately 2 months, making them a
valuable early indicator of market activity and prices.

The data is typically published as Excel or CSV files with monthly granularity
and complete provincial breakdown. Historical data covers multiple years.
"""

from __future__ import annotations

import io
import logging
import re
from typing import Any, Iterable

import openpyxl
import pandas as pd
from bs4 import BeautifulSoup

from ..framework.adapter import AdapterFailure, BaseAdapter, RunContext, SourceManifest
from ..framework.fetch import FetchPlan, RawPayload
from ..framework.record import CanonicalRecord, province

log = logging.getLogger(__name__)

# Try multiple possible URLs for CIEN data
CIEN_URLS = [
    "https://www.notariado.org/es-ES/notariado/estadisticas/cien",
    "https://www.notariado.org/cien",
    "https://www.notariado.org/es-ES/datos-estadisticos",
    "https://www.notariado.org/portal/estadisticas",
]

# CIEN's province names as they appear in Excel files (upper case, CIEN's spelling)
PROVINCE_NAMES = {
    "MADRID": "28",
    "BARCELONA": "08",
    "VALENCIA": "46",
    "MALAGA": "29",
    "SEVILLA": "41",
    "ILLES BALEARS": "07",
    "BIZKAIA": "48",
    "ZARAGOZA": "50",
}


def _num(value: Any) -> float | None:
    """Parse a value to float, handling Spanish number formats."""
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in ("nan", "null", "none", "_", "-"):
        return None
    try:
        # Handle both comma and period as decimal separators
        text = text.replace(".", "").replace(",", ".")
        return float(text)
    except ValueError:
        return None


class NotariadoAdapter(BaseAdapter):
    manifest = SourceManifest(
        source_id="notariado",
        publisher="Consejo General del Notariado",
        license="Reuse permitted (attribution)",
        attribution="Source: Consejo General del Notariado, CIEN",
        docs_url="https://www.notariado.org",
        cadence="monthly",
        geo_level="province",
        max_age_days=120,
        formats=("xlsx", "xls"),
        kind="official",
        min_rows=2,
        notes=(
            "CIEN (Centro de Información Estadística del Notariado) monthly housing "
            "sales and prices by province. Notaries record sales at signing, leading "
            "the land registry by approximately 2 months, so these statistics anticipate "
            "registry-based measures and are valuable early indicators of market activity."
        ),
    )

    def discover(self, ctx: RunContext) -> list[FetchPlan]:
        """Find the CIEN Excel file by trying multiple known URLs and patterns."""
        # Try to find the data from the listing pages
        for url in CIEN_URLS:
            try:
                listing_plan = FetchPlan(url=url, fmt="html", label=f"notariado-listing:{url[-30:]}")
                payload = ctx.fetcher.get(listing_plan)
                if payload is not None:
                    file_url = self._find_cien_link(payload.text())
                    if file_url:
                        return [FetchPlan(url=file_url, fmt="xlsx", label="cien-monthly")]
            except Exception as exc:  # noqa: BLE001
                log.debug("notariado: could not fetch %s: %s", url, exc)
                continue

        # If no file found via HTML parsing, try some direct URLs
        direct_urls = [
            "https://www.notariado.org/cien-estadisticas.xlsx",
            "https://www.notariado.org/datos/cien-series.xlsx",
            "https://www.notariado.org/estadisticas/cien-provincial.xlsx",
        ]
        for url in direct_urls:
            try:
                plan = FetchPlan(url=url, fmt="xlsx", label=f"cien-direct:{url[-30:]}", optional=True)
                payload = ctx.fetcher.get(plan)
                if payload is not None:
                    return [FetchPlan(url=url, fmt="xlsx", label="cien-monthly")]
            except Exception as exc:  # noqa: BLE001
                log.debug("notariado: direct URL %s failed: %s", url, exc)
                continue

        raise AdapterFailure(
            "could not locate CIEN (notary statistics) data. "
            "Tried multiple URLs on notariado.org; the data may be behind a login or the page layout may have changed."
        )

    @staticmethod
    def _find_cien_link(html: str) -> str | None:
        """Parse HTML to find the CIEN Excel file URL."""
        soup = BeautifulSoup(html, "lxml")
        for a in soup.find_all("a", href=True):
            href = a["href"]
            text = f"{href} {a.get_text()} {a.get('title', '')}".lower()
            # Look for links to Excel/XLS files with CIEN or estadística keywords
            if (
                (href.lower().endswith(".xls") or href.lower().endswith(".xlsx"))
                and any(kw in text for kw in ("cien", "estadístic", "estadistic", "provincial"))
            ):
                return href
        return None

    def parse(self, payload: RawPayload, ctx: RunContext) -> pd.DataFrame:
        """Load the Excel file and extract sales and price data."""
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(payload.content), data_only=True, read_only=True)
        except Exception as exc:  # noqa: BLE001
            raise AdapterFailure(f"failed to open Excel file: {exc}") from exc

        if not workbook.sheetnames:
            raise AdapterFailure("Excel file contains no sheets")

        rows = []
        # Try to find sheets with sales or price data
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_text = sheet_name.lower()

            # Skip metadata/summary sheets
            if any(kw in sheet_text for kw in ("resumen", "summary", "índice", "índex")):
                continue

            # Extract data from this sheet
            for row in sheet.iter_rows(min_row=1, values_only=True):
                if row and any(row):  # Skip empty rows
                    rows.append(row)

        if not rows:
            raise AdapterFailure("no data rows found in Excel file")

        df = pd.DataFrame(rows)
        return df

    def normalize(
        self, frame: pd.DataFrame, plan: FetchPlan, ctx: RunContext
    ) -> Iterable[CanonicalRecord]:
        """Convert DataFrame rows to canonical records for both sales and prices."""
        prov_codes = ctx.config.province_codes

        for row in frame.itertuples(index=False):
            if len(row) < 3:
                continue

            # Try to parse row with province name, year, month, sales_count, price_m2
            province_name = None
            year = None
            month = None
            sales_count = None
            price_m2 = None

            # Look for province name in first column
            if isinstance(row[0], str):
                prov_text = row[0].strip().upper()
                province_name = PROVINCE_NAMES.get(prov_text)

            if not province_name:
                continue

            # Try to extract numeric values
            numeric_vals = []
            for cell in row[1:]:
                n = _num(cell)
                if n is not None:
                    numeric_vals.append(n)

            if len(numeric_vals) < 3:
                continue

            # Common pattern: year, month, sales_count, price_m2
            # or: year, month, price_m2, sales_count
            if numeric_vals[0] > 1900 and numeric_vals[0] < 2100:
                year = int(numeric_vals[0])
                if 1 <= numeric_vals[1] <= 12:
                    month = int(numeric_vals[1])
                    # Remaining values are sales and price
                    if len(numeric_vals) >= 3:
                        sales_count = numeric_vals[2]
                    if len(numeric_vals) >= 4:
                        price_m2 = numeric_vals[3]
                else:
                    continue
            else:
                continue

            if not all([year, month, province_name]):
                continue

            # Emit sales record
            if sales_count is not None and sales_count >= 0:
                yield CanonicalRecord(
                    metric_id="notary_sales",
                    geo_id=province(province_name),
                    period=f"{year:04d}-{month:02d}",
                    value=sales_count,
                    unit="transactions",
                    source_id=self.manifest.source_id,
                )

            # Emit price record
            if price_m2 is not None and price_m2 >= 0:
                yield CanonicalRecord(
                    metric_id="notary_price_m2",
                    geo_id=province(province_name),
                    period=f"{year:04d}-{month:02d}",
                    value=price_m2,
                    unit="eur_m2",
                    source_id=self.manifest.source_id,
                )
