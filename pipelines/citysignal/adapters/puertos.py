"""Puertos del Estado — monthly national port-traffic summary.

Puertos del Estado (the state ports authority) publishes a "Resumen mensual
del tráfico portuario" for the whole state-managed port system, around the
25th of each month, in both XLSX and PDF at
``https://www.puertos.es/datos/estadisticas/mensuales``. That listing page
exposes a year selector (2012 onward) and, for each published month, a link
to the XLSX under ``/file-download/download/<private|public>/<id>``. There is
no query API and no bulk historical file with per-port monthly figures — the
"Estadística Histórica desde 1962" dashboard advertised on the site's
statistics landing page turned out, on inspection, to only serve national
tonnage yearbooks, not a per-port monthly passenger series — so ``discover``
does what ``BaseAdapter`` asks for: parse the listing page rather than
hard-code a deep link, for each of several recent years, and turn every XLSX
link it finds into a fetch plan.

**Reading the workbook.** Each monthly XLSX carries about 40 sheets, one per
traffic category; the one this adapter reads is ``Pasajeros total`` (a summed
total of regular-line ferry passengers and cruise passengers — exactly the
"cruise + ferry" figure this metric wants, so no manual summation is needed).
Its layout is dense and not row/column-labelled in a pandas-friendly way: a
title row, a blank row, a header row whose second cell names the report's own
month ("Junio "), a year row directly below it giving the two columns'
calendar years (e.g. 2025, 2026), then one row per Autoridad Portuaria with
that month's figure for the prior year and for the current year side by
side, ending at a "TOTAL" row. This adapter reads the month name and the two
years straight out of that header — never the URL or the listing page — so
the period a value is filed under is exactly the period the workbook itself
says it is.

Only the **current-year column** is taken from each file. Every calendar
month appears as the "current" column in exactly one monthly report (the one
published for that month) and later reappears as the "prior year" column in
the same month's report a year later; reading only "current" means every
(port, month) pair is emitted exactly once across however many files get
fetched, with no cross-file reconciliation needed.

**Geography: five ports, not six.** Puertos del Estado's Autoridad Portuaria
names match five of the eight cities' ports one-for-one: Barcelona, Valencia,
Málaga, Sevilla and Bilbao are each their own Autoridad Portuaria and each
report row matches a city name or alias exactly. Palma has no such row. Its
port is administered by the Autoridad Portuaria de Baleares, which reports
one combined figure spanning five ports across three islands (Palma,
Alcúdia, Maó, Eivissa, La Savina) — publishing that combined number under
``port-PMI`` would be exactly the geography substitution this project's
tooling exists to prevent (see ``record.py``'s module docstring). The
Balearic authority's own site, portsdebalears.com, advertises a per-port
statistics filter (Palma selectable on its own), but the filtered page
renders through client-side JavaScript with no backing data endpoint found
on inspection — it degrades to a static decorative image with no numbers in
the HTML at all, which is not something a scheduled pipeline can depend on.
``port_passengers`` for Palma is therefore left unpopulated rather than
filled with the Balearic-wide total; the gap is real and stays visible
instead of being papered over. This mirrors how ``mivau.py`` handles
``dwelling_stock`` being provincial rather than municipal in the source.
"""

from __future__ import annotations

import io
import re
import unicodedata
from typing import Any, Iterable

import openpyxl
import pandas as pd

from ..framework.adapter import AdapterFailure, BaseAdapter, RunContext, SourceManifest
from ..framework.fetch import FetchPlan, RawPayload
from ..framework.record import CanonicalRecord, port

LISTING_URL = "https://www.puertos.es/datos/estadisticas/mensuales"
FILE_URL = "https://www.puertos.es/file-download/download/{path}"
SHEET = "Pasajeros total"

# The listing page's year <select> is a 1-indexed offset from the current
# calendar year (value=1 is this year, value=2 is last year, ...). Six years
# is several years of monthly history without asking a site with no API for
# the full 2012-present run in one adapter run.
YEARS_BACK = 6

_XLSX_LINK = re.compile(
    r'file--mime-application-vnd-openxmlformats-officedocument-spreadsheetml-sheet'
    r'[^"]*"\s*>\s*<a href="([^"]+)"'
)

_MONTHS_ES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10, "noviembre": 11,
    "diciembre": 12,
}


def _strip_accents(text: str) -> str:
    return unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")


class PuertosAdapter(BaseAdapter):
    manifest = SourceManifest(
        source_id="puertos",
        publisher="Puertos del Estado",
        license="Reuse permitted (attribution)",
        attribution="Source: Puertos del Estado",
        docs_url="https://www.puertos.es/es-es/estadisticas",
        cadence="monthly",
        geo_level="port",
        max_age_days=120,
        formats=("xlsx",),
        kind="official",
        revisions_allowed=False,
        min_rows=0,
        notes=(
            "Monthly national port-traffic summary, 'Pasajeros total' sheet "
            "(cruise + regular-line ferry). Matches Barcelona, Valencia, Malaga, "
            "Sevilla and Bilbao one-for-one; Palma has no port-level breakdown in "
            "this source (Autoridad Portuaria de Baleares reports five ports as "
            "one combined figure) and is left unpopulated rather than promoted."
        ),
    )

    def discover(self, ctx: RunContext) -> list[FetchPlan]:
        import datetime as _dt

        current_year = _dt.date.today().year
        file_ids: dict[str, str] = {}  # id -> download path (private/public + id)

        for offset in range(YEARS_BACK):
            date_value = offset + 1  # 1 = current year, per the site's own <select>
            try:
                listing = ctx.fetcher.get(
                    FetchPlan(
                        url=LISTING_URL,
                        fmt="html",
                        label=f"listing:{current_year - offset}",
                        params={"date_value": date_value},
                    )
                )
            except Exception:  # noqa: BLE001 — one bad year must not sink the rest
                continue
            if listing is None:
                continue
            html = listing.text()
            for href in _XLSX_LINK.findall(html):
                match = re.search(r"/file-download/download/([a-z]+/\d+)", href)
                if match:
                    file_ids[match.group(1).rsplit("/", 1)[-1]] = match.group(1)

        if not file_ids:
            raise AdapterFailure("puertos: no monthly XLSX links found on any year's listing page")

        return [
            FetchPlan(
                url=FILE_URL.format(path=path),
                fmt="xlsx",
                label=f"resumen:{file_id}",
                # Interchangeable: one missing or malformed monthly report costs
                # that month, not the source. See module docstring.
                optional=True,
            )
            for file_id, path in sorted(file_ids.items())
        ]

    def parse(self, payload: RawPayload, ctx: RunContext) -> pd.DataFrame:
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(payload.content), data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise AdapterFailure(f"{payload.plan.label}: not a readable XLSX ({exc})") from exc
        if SHEET not in workbook.sheetnames:
            raise AdapterFailure(f"{payload.plan.label}: no {SHEET!r} sheet")
        sheet = workbook[SHEET]
        rows = list(sheet.iter_rows(values_only=True))

        header_idx = next(
            (i for i, row in enumerate(rows) if row and str(row[0]).strip() == "Autoridad Portuaria"),
            None,
        )
        if header_idx is None or header_idx + 1 >= len(rows):
            raise AdapterFailure(f"{payload.plan.label}: could not find the header row")

        month_name = _strip_accents(str(rows[header_idx][1] or "").strip().lower())
        month = _MONTHS_ES.get(month_name)
        year_row = rows[header_idx + 1]
        year_cur = year_row[2] if len(year_row) > 2 else None
        if month is None or year_cur is None:
            raise AdapterFailure(f"{payload.plan.label}: unreadable month/year header")

        records: list[dict[str, Any]] = []
        for row in rows[header_idx + 2 :]:
            if not row or row[0] is None:
                continue
            name = str(row[0]).strip()
            if name == "TOTAL" or name.startswith("Incluye"):
                break
            value = row[2] if len(row) > 2 else None
            if value is None:
                continue
            records.append({"autoridad": name, "value": float(value)})

        if not records:
            raise AdapterFailure(f"{payload.plan.label}: no port rows under the header")

        frame = pd.DataFrame(records)
        frame["period"] = f"{int(year_cur):04d}-{month:02d}"
        return frame

    def normalize(
        self, frame: pd.DataFrame, plan: FetchPlan, ctx: RunContext
    ) -> Iterable[CanonicalRecord]:
        # Cities whose ports map one-for-one to a named Autoridad Portuaria.
        # Palma naturally falls out here: neither "Palma" nor "Palma de
        # Mallorca" nor any alias equals "Baleares", so it never matches — see
        # module docstring for why that is the honest outcome, not a bug.
        city_by_ap_name: dict[str, Any] = {}
        for city in ctx.config.cities:
            if not city.ports:
                continue
            for candidate in (city.name, *city.aliases):
                city_by_ap_name[candidate] = city

        out: list[CanonicalRecord] = []
        for row in frame.itertuples(index=False):
            city = city_by_ap_name.get(row.autoridad)
            if city is None:
                continue
            out.append(
                CanonicalRecord(
                    metric_id="port_passengers",
                    geo_id=port(city.ports[0]),
                    period=row.period,
                    value=row.value,
                    unit="passengers",
                    source_id=self.manifest.source_id,
                )
            )
        return out
